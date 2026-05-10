"""
【モーキス】制作進行表からタスク期限を抽出してDiscordに通知するスクリプト

使い方:
  - 環境変数 DISCORD_WEBHOOK_URL に Webhook URL を設定
  - 設定なしで実行するとコンソールに出力（ドライラン）

依存:
  pip install openpyxl requests
"""
import os
import sys
import io
import json
from datetime import datetime, timedelta, date
from urllib import request as urlrequest

import openpyxl

# Windows コンソールの文字化け対策
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXCEL_PATH = os.environ.get(
    "PROGRESS_XLSX",
    r"C:\Users\kohet\モーキス作業ディレクトリ\【モーキス】制作進行表.xlsx",
)
WEBHOOK_URL = os.environ.get("DISCORD_WEBHOOK_URL", "")

# シートごとの構造（header_row は 0 始まり）
SHEET_CONFIG = {
    "脚本":         {"header_row": 1},
    "音楽・SE":     {"header_row": 1},
    "撮影":         {"header_row": 1},
    "プログラム":   {"header_row": 2},
    "グラフィック": {"header_row": 1},
    "デザイン":     {"header_row": 1},
    "広報":         {"header_row": 1},
    "スクリプト":   {"header_row": 1},
}

# 完了扱い・対象外のステータス
EXCLUDE_STATUS = {"制作完了", "取消"}


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def collect_tasks():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    tasks = []

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = cfg["header_row"]

        # ヘッダー行から列インデックスを動的に取得
        headers = [c.value for c in ws[header_row + 1]]
        col_map = {}
        for i, h in enumerate(headers):
            if h in ("大項目", "制作物名", "締切", "担当者", "進捗"):
                col_map[h] = i

        required = {"制作物名", "締切", "進捗"}
        if not required.issubset(col_map):
            continue

        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            name = row[col_map["制作物名"]] if col_map.get("制作物名") is not None else None
            deadline = to_date(row[col_map["締切"]]) if col_map.get("締切") is not None else None
            status = row[col_map["進捗"]] if col_map.get("進捗") is not None else None
            assignee = row[col_map.get("担当者")] if col_map.get("担当者") is not None else None
            category = row[col_map.get("大項目")] if col_map.get("大項目") is not None else None

            if not name or deadline is None:
                continue
            if status in EXCLUDE_STATUS:
                continue

            tasks.append({
                "section": sheet_name,
                "category": category or "",
                "name": str(name),
                "deadline": deadline,
                "assignee": assignee or "未割当",
                "status": status or "未設定",
            })
    return tasks


def categorize(tasks, base_date):
    tomorrow = base_date + timedelta(days=1)
    week_end = base_date + timedelta(days=8)  # 8日後の前日まで = 1週間以内

    buckets = {"overdue": [], "today": [], "tomorrow": [], "week": []}
    for t in tasks:
        d = t["deadline"]
        if d < base_date:
            buckets["overdue"].append(t)
        elif d == base_date:
            buckets["today"].append(t)
        elif d == tomorrow:
            buckets["tomorrow"].append(t)
        elif tomorrow < d < week_end:
            buckets["week"].append(t)

    for arr in buckets.values():
        arr.sort(key=lambda t: (t["deadline"], t["section"]))
    return buckets


def format_task(t):
    cat = f"/{t['category']}" if t["category"] else ""
    return f"・`{t['section']}{cat}` **{t['name']}** — {t['assignee']}（{t['status']}）期限: {t['deadline'].strftime('%m/%d')}"


def build_message(buckets, base_date):
    lines = [
        f"📋 **【モーキス】制作進行 — タスク期限通知**（{base_date.strftime('%Y-%m-%d')} 時点）",
        "",
    ]

    sections = [
        ("🔴 **本日中が期限**", "today"),
        ("🟡 **明日中が期限**", "tomorrow"),
        ("🟢 **1週間以内が期限**", "week"),
    ]
    for label, key in sections:
        lines.append(label)
        if buckets[key]:
            lines.extend(format_task(t) for t in buckets[key])
        else:
            lines.append("_該当なし_")
        lines.append("")

    if buckets["overdue"]:
        lines.append(f"⚠️ **期限切れ・未完了（{len(buckets['overdue'])}件）**")
        # 期限切れは多すぎる可能性があるので最新10件のみ
        for t in buckets["overdue"][-10:]:
            lines.append(format_task(t))
        if len(buckets["overdue"]) > 10:
            lines.append(f"_…他 {len(buckets['overdue']) - 10}件_")

    return "\n".join(lines)


def split_for_discord(content, limit=1900):
    """Discord の 2000 文字制限に合わせて分割"""
    if len(content) <= limit:
        return [content]
    chunks, current = [], ""
    for line in content.split("\n"):
        if len(current) + len(line) + 1 > limit:
            chunks.append(current)
            current = line
        else:
            current = current + "\n" + line if current else line
    if current:
        chunks.append(current)
    return chunks


def send_discord(content):
    for chunk in split_for_discord(content):
        data = json.dumps({"content": chunk}).encode("utf-8")
        req = urlrequest.Request(
            WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        urlrequest.urlopen(req).read()


def main():
    base_date = date.today()
    tasks = collect_tasks()
    buckets = categorize(tasks, base_date)
    message = build_message(buckets, base_date)

    print(f"[INFO] 対象タスク数: {len(tasks)}")
    print(f"[INFO] 本日: {len(buckets['today'])}件 / 明日: {len(buckets['tomorrow'])}件 / 1週間以内: {len(buckets['week'])}件 / 期限切れ: {len(buckets['overdue'])}件")
    print("---")
    print(message)
    print("---")

    if WEBHOOK_URL:
        send_discord(message)
        print("[INFO] Discordに送信しました")
    else:
        print("[INFO] DISCORD_WEBHOOK_URL 未設定のためコンソール出力のみ")


if __name__ == "__main__":
    main()
